import openpyxl
from bs4 import BeautifulSoup
import urllib2
from time import sleep, localtime, strftime

#method to retrieve price from url
def retrieve_price(url):
	content = urllib2.urlopen(url).read()
	soup = BeautifulSoup(content, 'html.parser')
	#get this from original webscraping script
	price = soup.find('span', {'class':'Trsdu(0.3s) Fw(b) Fz(36px) Mb(-4px) D(ib)'}).string  
	price = float(price) #convert from unicode to string
	return price


while True:
        try:
                #open Savings file
                savings_workbook = openpyxl.load_workbook('/Users/chowjiaying/Desktop/Finances/Savings.xlsx')
                #open Portfolio worksheet
                portfolio = savings_workbook.get_sheet_by_name('Portfolio 2020')
                

                start_column = 2
                start_row = 2

                #loop to update stock prices of all stocks in portfolio
                current_cell = portfolio.cell(row=start_row, column=start_column)
                while current_cell.value is not None:
                        if current_cell.value != 'SOLD':
                                url = current_cell.value
                                #go to url to retrieve price
                                price = retrieve_price(url)
                                #update stock price value
                                portfolio.cell(row=start_row+2, column=start_column+5, value=price)


                        start_row += 4
                        current_cell = portfolio.cell(row=start_row, column=start_column)



        except:
                current_time = strftime("%I:%M:%S %p", localtime())
                print("An error occured at {}!".format(current_time)) #Don't save anything, figure out logger later
		print(e)
                sleep(3600) # try an hour later
                continue
        else:
                current_time = strftime("%I:%M:%S %p", localtime())
                savings_workbook.save('/Users/chowjiaying/Desktop/Finances/Savings.xlsx') #override old workbook
                print("Success in updating stock prices at {}!".format(current_time))
                break
