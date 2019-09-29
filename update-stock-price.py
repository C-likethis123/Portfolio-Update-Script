import openpyxl
from bs4 import BeautifulSoup
import urllib2

#TODO: replace $NAME_OF_EXCEL_FILE with the file path of your portfolio
 
 
#method to retrieve price from url
def retrieve_price(url):
	content = urllib2.urlopen(url).read()
	soup = BeautifulSoup(content, 'html.parser')
	#get this from original webscraping script
	price = soup.find('span', {'class':'Trsdu(0.3s)'}).string
	price = float(price) #convert from unicode to string
	return price


try:
	#open Savings file
	savings_workbook = openpyxl.load_workbook('$NAME_OF_EXCEL_FILE')
	#open Portfolio worksheet
	portfolio = savings_workbook.get_sheet_by_name('Portfolio 2019')
	

	start_column = 2
	start_row = 2

	#loop to update stock prices of all stocks in portfolio
	current_cell = portfolio.cell(row=start_row, column=start_column)
	while current_cell.value is not None:
		if current_cell.value != 'SOLD':
			url = current_cell.value
			#go to url to retrieve price
			price = retrieve_price(url)
			#update stock price value
			portfolio.cell(row=start_row+2, column=start_column+5, value=price)


		start_row += 4
		current_cell = portfolio.cell(row=start_row, column=start_column)



except:
	print("An error occured!") #Don't save anything, figure out logger later
else:
	savings_workbook.save('$NAME_OF_EXCEL_FILE') #override old workbook
	print("Success in updating stock prices!")
